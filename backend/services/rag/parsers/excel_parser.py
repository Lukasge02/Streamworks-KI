"""
Enterprise Excel & CSV Parser for Parameter Extraction

Supports:
- .xlsx (OpenXML Excel)
- .xls (Legacy Excel)
- .csv (Comma/Semicolon separated)

Features:
- Automatic header detection
- Parameter mapping via ParameterRegistry
- Multi-stream extraction (one row = one stream)
- Smart delimiter detection for CSV
"""

import io
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, field

from .base_parser import BaseParser, ParsedDocument, DocumentType


@dataclass
class ExtractedRow:
    """A single row of extracted parameters from spreadsheet"""
    row_number: int
    raw_data: Dict[str, Any]
    mapped_params: Dict[str, str]
    confidence: float = 0.8
    warnings: List[str] = field(default_factory=list)


@dataclass
class SpreadsheetExtractionResult:
    """Result of extracting parameters from spreadsheet"""
    rows: List[ExtractedRow]
    header_mapping: Dict[str, str]  # Original header -> StreamworksParam key
    unmapped_columns: List[str]
    total_rows: int
    sheet_name: Optional[str] = None


def _get_header_mappings() -> Dict[str, str]:
    """
    Get header mappings from ParameterRegistry.
    Falls back to basic mappings if registry unavailable.
    """
    try:
        from services.ai.parameter_registry import get_parameter_registry
        registry = get_parameter_registry()
        return registry.get_excel_mappings()
    except Exception as e:
        print(f"⚠️ Could not load ParameterRegistry: {e}, using fallback mappings")
        # Fallback basic mappings
        return {
            "streamname": "stream_name",
            "name": "stream_name",
            "sourceagent": "source_agent",
            "source": "source_agent",
            "targetagent": "target_agent",
            "target": "target_agent",
            "description": "short_description",
            "schedule": "schedule",
            "starttime": "start_time",
            "priority": "stream_priority",
        }


class ExcelParser(BaseParser):
    """
    Enterprise Excel/CSV Parser for Parameter Extraction
    
    Extracts stream parameters from:
    - Excel files (.xlsx, .xls)
    - CSV files (auto-detects delimiter)
    
    Each row in the spreadsheet represents one stream configuration.
    Headers are automatically mapped to StreamworksParams fields via ParameterRegistry.
    """
    
    def __init__(self):
        self._openpyxl = None
        self._xlrd = None
        self._header_mappings = None
    
    @property
    def header_mappings(self) -> Dict[str, str]:
        """Lazy load header mappings from ParameterRegistry"""
        if self._header_mappings is None:
            self._header_mappings = _get_header_mappings()
        return self._header_mappings
    
    @property
    def openpyxl(self):
        """Lazy load openpyxl"""
        if self._openpyxl is None:
            try:
                import openpyxl
                self._openpyxl = openpyxl
            except ImportError as e:
                raise ImportError(
                    "openpyxl not installed. Run: pip install openpyxl"
                ) from e
        return self._openpyxl
    
    @property
    def supported_types(self) -> List[DocumentType]:
        return [DocumentType.EXCEL, DocumentType.CSV]
    
    @property
    def name(self) -> str:
        return "ExcelParser"
    
    def _detect_type(self, filename: str) -> DocumentType:
        """Override to handle Excel/CSV detection"""
        ext = filename.lower().split(".")[-1] if "." in filename else ""
        if ext in ("xlsx", "xls"):
            return DocumentType.EXCEL
        elif ext == "csv":
            return DocumentType.CSV
        return super()._detect_type(filename)
    
    def parse(self, content: bytes, filename: str, **kwargs) -> ParsedDocument:
        """
        Parse Excel/CSV file and return standardized document
        
        For parameter extraction, use parse_for_parameters() instead.
        """
        doc_type = self._detect_type(filename)
        
        if doc_type == DocumentType.CSV:
            text_content, tables = self._parse_csv(content)
        else:
            text_content, tables = self._parse_excel(content)
        
        return ParsedDocument(
            content=text_content,
            filename=filename,
            doc_type=doc_type,
            title=filename.rsplit(".", 1)[0],
            tables=tables,
            metadata={
                "source": filename,
                "parsing_engine": "excel_parser",
                "table_count": len(tables),
            },
            word_count=len(text_content.split()),
            parsing_method="excel_parser",
        )
    
    def parse_for_parameters(
        self, 
        content: bytes, 
        filename: str,
        sheet_name: Optional[str] = None,
    ) -> SpreadsheetExtractionResult:
        """
        Extract stream parameters from spreadsheet
        
        Args:
            content: Raw file bytes
            filename: Original filename
            sheet_name: Specific sheet to parse (Excel only)
            
        Returns:
            SpreadsheetExtractionResult with extracted parameter rows
        """
        doc_type = self._detect_type(filename)
        
        if doc_type == DocumentType.CSV:
            return self._extract_params_from_csv(content)
        else:
            return self._extract_params_from_excel(content, sheet_name)
    
    def _parse_excel(self, content: bytes) -> Tuple[str, List[Dict]]:
        """Parse Excel file to text and tables"""
        workbook = self.openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        
        all_text = []
        tables = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_text = [f"=== Sheet: {sheet_name} ==="]
            
            rows_data = []
            headers = []
            
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                # Skip completely empty rows
                if all(cell is None for cell in row):
                    continue
                    
                row_values = [str(cell) if cell is not None else "" for cell in row]
                
                if row_idx == 0:
                    headers = row_values
                else:
                    rows_data.append(row_values)
                
                sheet_text.append(" | ".join(row_values))
            
            if headers:
                tables.append({
                    "sheet": sheet_name,
                    "headers": headers,
                    "rows": rows_data,
                })
            
            all_text.extend(sheet_text)
        
        workbook.close()
        return "\n".join(all_text), tables
    
    def _parse_csv(self, content: bytes) -> Tuple[str, List[Dict]]:
        """Parse CSV file to text and tables"""
        import csv
        
        # Decode content
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        
        # Detect delimiter
        delimiter = self._detect_csv_delimiter(text)
        
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows = list(reader)
        
        if not rows:
            return "", []
        
        headers = rows[0]
        data_rows = rows[1:]
        
        text_content = "\n".join(" | ".join(row) for row in rows)
        
        tables = [{
            "sheet": "CSV",
            "headers": headers,
            "rows": data_rows,
        }]
        
        return text_content, tables
    
    def _detect_csv_delimiter(self, text: str) -> str:
        """Auto-detect CSV delimiter (comma, semicolon, tab)"""
        first_lines = "\n".join(text.split("\n")[:5])
        
        delimiters = {
            ",": first_lines.count(","),
            ";": first_lines.count(";"),
            "\t": first_lines.count("\t"),
        }
        
        # Return the most common delimiter
        return max(delimiters, key=delimiters.get)
    
    def _extract_params_from_excel(
        self, 
        content: bytes, 
        sheet_name: Optional[str] = None
    ) -> SpreadsheetExtractionResult:
        """Extract parameters from Excel file"""
        workbook = self.openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        
        # Select sheet
        if sheet_name and sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
            sheet_name = sheet.title
        
        # Read all rows
        all_rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
        
        if not all_rows:
            return SpreadsheetExtractionResult(
                rows=[],
                header_mapping={},
                unmapped_columns=[],
                total_rows=0,
                sheet_name=sheet_name,
            )
        
        # First row is headers
        raw_headers = [str(h).strip() if h else "" for h in all_rows[0]]
        
        # Map headers
        header_mapping, unmapped = self._map_headers(raw_headers)
        
        # Extract data rows
        extracted_rows = []
        for row_idx, row in enumerate(all_rows[1:], start=2):
            if all(cell is None for cell in row):
                continue
            
            raw_data = {}
            mapped_params = {}
            
            for col_idx, cell in enumerate(row):
                if col_idx < len(raw_headers):
                    header = raw_headers[col_idx]
                    value = str(cell).strip() if cell is not None else ""
                    
                    raw_data[header] = value
                    
                    # Map to parameter if we have a mapping
                    normalized = header.lower().replace(" ", "_").replace("-", "_")
                    if normalized in self.header_mappings:
                        param_key = self.header_mappings[normalized]
                        if value:
                            mapped_params[param_key] = value
                    elif header in header_mapping:
                        param_key = header_mapping[header]
                        if value:
                            mapped_params[param_key] = value
            
            # Only include rows with at least one mapped parameter
            if mapped_params:
                extracted_rows.append(ExtractedRow(
                    row_number=row_idx,
                    raw_data=raw_data,
                    mapped_params=mapped_params,
                    confidence=self._calculate_row_confidence(mapped_params),
                ))
        
        return SpreadsheetExtractionResult(
            rows=extracted_rows,
            header_mapping=header_mapping,
            unmapped_columns=unmapped,
            total_rows=len(extracted_rows),
            sheet_name=sheet_name,
        )
    
    def _extract_params_from_csv(self, content: bytes) -> SpreadsheetExtractionResult:
        """Extract parameters from CSV file"""
        import csv
        
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        
        delimiter = self._detect_csv_delimiter(text)
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows = list(reader)
        
        if not rows:
            return SpreadsheetExtractionResult(
                rows=[],
                header_mapping={},
                unmapped_columns=[],
                total_rows=0,
            )
        
        raw_headers = [h.strip() for h in rows[0]]
        header_mapping, unmapped = self._map_headers(raw_headers)
        
        extracted_rows = []
        for row_idx, row in enumerate(rows[1:], start=2):
            if not any(cell.strip() for cell in row):
                continue
            
            raw_data = {}
            mapped_params = {}
            
            for col_idx, cell in enumerate(row):
                if col_idx < len(raw_headers):
                    header = raw_headers[col_idx]
                    value = cell.strip()
                    
                    raw_data[header] = value
                    
                    normalized_header = header.lower().replace(" ", "_").replace("-", "_")
                    if normalized_header in self.header_mappings:
                        param_key = self.header_mappings[normalized_header]
                        if value:
                            mapped_params[param_key] = value
            
            if mapped_params:
                extracted_rows.append(ExtractedRow(
                    row_number=row_idx,
                    raw_data=raw_data,
                    mapped_params=mapped_params,
                    confidence=self._calculate_row_confidence(mapped_params),
                ))
        
        return SpreadsheetExtractionResult(
            rows=extracted_rows,
            header_mapping=header_mapping,
            unmapped_columns=unmapped,
            total_rows=len(extracted_rows),
        )
    
    def _map_headers(self, headers: List[str]) -> Tuple[Dict[str, str], List[str]]:
        """Map raw headers to StreamworksParams field names"""
        mapping = {}
        unmapped = []
        
        for header in headers:
            if not header:
                continue
                
            normalized = header.lower().replace(" ", "_").replace("-", "_")
            
            if normalized in self.header_mappings:
                mapping[header] = self.header_mappings[normalized]
            else:
                unmapped.append(header)
        
        return mapping, unmapped
    
    def _calculate_row_confidence(self, params: Dict[str, str]) -> float:
        """Calculate confidence score for extracted row"""
        critical_params = ["stream_name", "source_agent", "target_agent", "agent_detail"]
        high_params = ["short_description", "schedule", "start_time"]
        
        score = 0.5  # Base score
        
        # Boost for critical params
        for param in critical_params:
            if param in params and params[param]:
                score += 0.1
        
        # Smaller boost for high params
        for param in high_params:
            if param in params and params[param]:
                score += 0.05
        
        return min(score, 1.0)
