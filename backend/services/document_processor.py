"""
Document processing pipeline: parse, chunk, embed, store.

Supports PDF, DOCX, XLSX, and plain text files.
Stores the original file in MinIO and indexed chunks in Qdrant.
"""

import logging
import uuid
from services import vector_store, file_storage

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def _parse_file(filename: str, file_bytes: bytes, content_type: str) -> str:
    """
    Extract plain text from a file based on its content type.

    Args:
        filename: Original filename (used for extension fallback).
        file_bytes: Raw bytes of the uploaded file.
        content_type: MIME type string.

    Returns:
        Extracted text content as a single string.

    Raises:
        ValueError: If the file type is not supported.
    """
    lower_name = filename.lower()

    # PDF
    if content_type == "application/pdf" or lower_name.endswith(".pdf"):
        return _parse_pdf(file_bytes)

    # DOCX
    if (
        content_type
        == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        or lower_name.endswith(".docx")
    ):
        return _parse_docx(file_bytes)

    # XLSX
    if (
        content_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        or lower_name.endswith(".xlsx")
    ):
        return _parse_xlsx(file_bytes)

    # Plain text (txt, csv, md, json, xml, yaml, etc.)
    if content_type.startswith("text/") or lower_name.endswith(
        (".txt", ".csv", ".md", ".json", ".xml", ".yaml", ".yml")
    ):
        return file_bytes.decode("utf-8", errors="replace")

    raise ValueError(
        f"Unsupported file type: content_type={content_type}, filename={filename}"
    )


def _parse_pdf(file_bytes: bytes) -> str:
    """Extract text from a PDF using PyMuPDF (fitz)."""
    import fitz  # PyMuPDF

    pages: list[str] = []
    with fitz.open(stream=file_bytes, filetype="pdf") as doc:
        for page in doc:
            text = page.get_text()
            if text.strip():
                pages.append(text)
    return "\n\n".join(pages)


def _parse_docx(file_bytes: bytes) -> str:
    """Extract text from a DOCX file using python-docx."""
    import io
    from docx import Document

    doc = Document(io.BytesIO(file_bytes))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def _parse_xlsx(file_bytes: bytes) -> str:
    """Extract text from an XLSX file using openpyxl."""
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"--- Sheet: {sheet_name} ---")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            line = " | ".join(cells).strip()
            if line and line != "|":
                parts.append(line)

    wb.close()
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Chunking
# ---------------------------------------------------------------------------

def _chunk_text(
    text: str,
    chunk_size: int = 800,
    overlap: int = 200,
) -> list[str]:
    """
    Split text into overlapping chunks, splitting on paragraph boundaries.

    The algorithm first splits text into paragraphs (double newlines), then
    greedily groups paragraphs into chunks that stay within ``chunk_size``
    characters. If a single paragraph exceeds ``chunk_size``, it is split
    on sentence boundaries or, as a last resort, by character count.

    Consecutive chunks share approximately ``overlap`` characters to
    preserve context at boundaries.

    Args:
        text: The full text to chunk.
        chunk_size: Target maximum characters per chunk.
        overlap: Approximate character overlap between consecutive chunks.

    Returns:
        A list of non-empty text chunks.
    """
    if not text.strip():
        return []

    paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]

    # If there are no paragraph breaks, fall back to single newline splits
    if len(paragraphs) <= 1 and len(text) > chunk_size:
        paragraphs = [p.strip() for p in text.split("\n") if p.strip()]

    chunks: list[str] = []
    current_chunk = ""

    for para in paragraphs:
        # If adding this paragraph would exceed chunk_size, flush current chunk
        if current_chunk and len(current_chunk) + len(para) + 2 > chunk_size:
            chunks.append(current_chunk.strip())
            # Start new chunk with overlap from the end of the previous chunk
            if overlap > 0 and len(current_chunk) > overlap:
                current_chunk = current_chunk[-overlap:] + "\n\n" + para
            else:
                current_chunk = para
        else:
            if current_chunk:
                current_chunk += "\n\n" + para
            else:
                current_chunk = para

        # Handle single paragraphs that are larger than chunk_size
        while len(current_chunk) > chunk_size * 1.5:
            split_point = current_chunk.rfind(". ", 0, chunk_size)
            if split_point == -1:
                split_point = chunk_size
            else:
                split_point += 1  # Include the period

            chunks.append(current_chunk[:split_point].strip())
            remaining = current_chunk[split_point:].strip()
            if overlap > 0 and len(current_chunk[:split_point]) > overlap:
                current_chunk = current_chunk[split_point - overlap :].strip()
            else:
                current_chunk = remaining

    if current_chunk.strip():
        chunks.append(current_chunk.strip())

    return chunks


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def process_document(
    filename: str,
    file_bytes: bytes,
    content_type: str,
) -> dict:
    """
    Full document ingestion pipeline:

    1. Parse the file to extract plain text.
    2. Chunk the text with overlap.
    3. Embed all chunks via OpenAI.
    4. Upsert chunks with embeddings into Qdrant.
    5. Store the original file in MinIO.

    Args:
        filename: Original filename.
        file_bytes: Raw file bytes.
        content_type: MIME type of the file.

    Returns:
        A dict with document_id, filename, and chunks_count.
    """
    document_id = str(uuid.uuid4())
    logger.info(
        "Processing document: %s (id=%s, type=%s, size=%d bytes)",
        filename,
        document_id,
        content_type,
        len(file_bytes),
    )

    # 1. Parse
    text = _parse_file(filename, file_bytes, content_type)
    if not text.strip():
        logger.warning("No text extracted from %s", filename)
        return {
            "document_id": document_id,
            "filename": filename,
            "chunks_count": 0,
        }

    logger.info("Extracted %d characters from %s", len(text), filename)

    # 2. Chunk
    chunks_text = _chunk_text(text, chunk_size=800, overlap=200)
    logger.info("Created %d chunks from %s", len(chunks_text), filename)

    # 3. Embed
    embeddings = vector_store.embed_texts(chunks_text)

    # 4. Upsert to Qdrant
    chunk_records = []
    for idx, (chunk_text_item, embedding) in enumerate(
        zip(chunks_text, embeddings)
    ):
        chunk_records.append({
            "text": chunk_text_item,
            "embedding": embedding,
            "metadata": {
                "document_name": filename,
                "chunk_index": idx,
                "page": idx + 1,  # Approximate page mapping
            },
        })

    vector_store.ensure_collection()
    vector_store.upsert_chunks(document_id, chunk_records)

    # 5. Store original in MinIO
    object_name = f"{document_id}/{filename}"
    file_storage.upload_file(object_name, file_bytes, content_type)

    logger.info(
        "Document processing complete: %s (%d chunks)", filename, len(chunks_text)
    )

    return {
        "document_id": document_id,
        "filename": filename,
        "chunks_count": len(chunks_text),
    }
